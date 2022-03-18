from openpyxl import open, Workbook
from openpyxl import styles
import re


class Excel:
    def __init__(self, path=None, read_only=True):
        self.initialData = ()
        self.structuredData = ()
        self.path = path
        if read_only and path:
            self.file = open(path, read_only=read_only)
        else:
            self.file = Workbook()
        self.file.active.title = 'Sheet_1'
        self.line_number = 2

    def auto_size_cols(self):
        for col in self.file.active.columns:
            max_length = max(map(lambda x: len(x.value or ''), col))
            column_letter = col[-1].column_letter  # Get the column name
            new_width = (max_length + 1) * 1.2
            self.file.active.column_dimensions[column_letter].width = new_width

    def output_values(self, output_dict: dict):
        values = self.rec_find_vals_keys(output_dict)['vals']

        next_row = self.file.active.max_row + 1
        my_range = self.file.active.iter_cols(min_row=next_row, max_row=next_row, max_col=len(values), min_col=1)
        for cell, val in zip(my_range, values):
            cell[0].value = val
            cell[0].border = styles.Border(right=styles.Side(style='thin'),
                                           bottom=styles.Side(style='thin'))

    def output_col_names(self, output_dict: dict):
        keys = list(output_dict.keys())
        vals, keys_with_dict = self.rec_find_vals_keys(output_dict).values()

        next_row = self.file.active.max_row + 1
        my_range = list(self.file.active.iter_cols(min_row=next_row, max_row=next_row, max_col=len(vals), min_col=1))
        cell = key = 0
        while cell != len(my_range) and key != len(keys):
            my_range[cell][0].value = keys[key]
            my_range[cell][0].font = styles.Font(bold=True)
            my_range[cell][0].alignment = styles.Alignment(horizontal='center', vertical='center')
            my_range[cell][0].border = styles.Border(left=styles.Side(style='thick'),
                                                     right=styles.Side(style='thick'),
                                                     top=styles.Side(style='thick'),
                                                     bottom=styles.Side(style='thick'))
            if keys[key] in keys_with_dict:
                self.file.active.insert_cols(idx=my_range[cell][0].column + 1,
                                             amount=len(keys_with_dict[keys[key]]) - 1)

                self.file.active.merge_cells(start_row=my_range[cell][0].row,
                                             start_column=my_range[cell][0].column,
                                             end_row=my_range[cell][0].row,
                                             end_column=my_range[cell][0].column + len(keys_with_dict[keys[key]]) - 1)

                sub_range = self.file.active.iter_cols(min_row=my_range[cell][0].row + 1,
                                                       max_row=my_range[cell][0].row + 1,
                                                       min_col=my_range[cell][0].column,
                                                       max_col=my_range[cell][0].column + len(
                                                           keys_with_dict[keys[key]]) - 1)
                for sub_cell, val in zip(sub_range, keys_with_dict[keys[key]]):
                    sub_cell[0].value = val
                    sub_cell[0].font = styles.Font(italic=True)
                    sub_cell[0].alignment = styles.Alignment(horizontal='center', vertical='center')
                    sub_cell[0].border = styles.Border(left=styles.Side(style='thick'),
                                                       right=styles.Side(style='thick'),
                                                       top=styles.Side(style='thick'),
                                                       bottom=styles.Side(style='thick'))

                cell += len(keys_with_dict[keys[key]]) - 2 or 1
                key += 1
            else:
                self.file.active.merge_cells(start_row=my_range[cell][0].row,
                                             start_column=my_range[cell][0].column,
                                             end_column=my_range[cell][0].column,
                                             end_row=my_range[cell][0].row + 1)
                cell += 1
                key += 1

    def output_in_cell(self, value, cell=None):
        if not cell:
            self.__output_name(value)
        else:
            self.file.active[cell] = value

    def __output_name(self, value):
        # index = self.file.active.max_row + 1
        self.file.active["A1"] = value
        # self.file.active.merge_cells(f"A{index}:C{index}")
        self.file.active.merge_cells(start_row=1,
                                     start_column=1,
                                     end_row=1,
                                     end_column=self.file.active.max_column)
        self.file.active["A1"].font = styles.Font(bold=True, size=12, color="000066CC")
        self.file.active["A1"].alignment = styles.Alignment(horizontal='center', vertical='center')

    def new_sheet(self):
        new_index = len(self.file.get_sheet_names()) + 1
        self.file.create_sheet(f"Sheet_{new_index}")
        self.file.active = new_index - 1
        # self.file = self.file[f"Sheet_{new_index}"]

    def count_sheets(self):
        return len(self.file.get_sheet_names())

    def save(self, path=None):
        if path:
            self.file.save(path)
        self.file.save(self.path)

    # recursion find values and keys with dict-values in dict
    @staticmethod
    def rec_find_vals_keys(a, vals=None, keys_with_dict=None):
        if keys_with_dict is None:
            keys_with_dict = {}
        if vals is None:
            vals = []

        for key in a:
            if isinstance(a[key], dict):
                keys_with_dict.get(key, keys_with_dict.setdefault(key, list(a[key].keys())))
                # keys_with_dict.append({key: list(a[key].keys())})
                vals = Excel().rec_find_vals_keys(a[key], vals, keys_with_dict)['vals']
            else:
                vals.append(a[key])
        return {'vals': vals, 'keys_with_dict': keys_with_dict}

    @property
    def num_goods(self):
        return self.file.active.max_row - 2

    def readline(self):
        self.initialData += (self.file.active[self.line_number][1].value[2:],)
        self.line_number += 1
        return self

    # def importgoods(self):
    #     sheet = self.file.active
    #     self.initialData = ()
    #     for row in range(2, sheet.max_row):
    #         self.initialData += (sheet[row][1].value[2:],)

    def close_file(self):
        self.file.close()

    def structurizedata(self, params1, params2):
        # Выделение параметра и значения в каждой паре
        def pars(s):
            spl = re.split(r'' + restr, s)[1:]
            return spl[0].lower(), spl[1].strip().lower()

        # Генерирует выбранный порядок параметров для поискового запроса
        def addtosearch(*params):
            search = ()
            for temp in params:
                if temp in structName:
                    if temp == 'головка' or temp == 'класс прочности':
                        search += (temp, structName[temp])
                    else:
                        search += (structName[temp],)
            return search

        rawName = self.initialData[self.line_number - 3]

        structName = {'исходная строка': rawName}
        if rawName.find(';') == -1:  # Описание товара 1 типа
            restr1 = '(' + '|'.join(params1.values()) + ')'
            spl = re.split(r'' + restr1, rawName)
            if spl[1] == 'головка':
                obozn = spl[2].split()[1]
                structName['обозначение'] = obozn
                spl[2] = spl[2].replace(obozn, '')
                structName['название'] = spl[0]
            else:
                t = spl[0].split()
                structName['обозначение'] = t[-1]
                structName['название'] = ' '.join(t[:-1])
            for key, value in params1.items():
                if value in spl:
                    i = spl.index(value)
                    if key == 'стандарт':
                        structName[key] = (spl[i] + spl[i + 1]).strip()
                    else:
                        structName[key] = spl[i + 1].strip()
            structName['поиск'] = structName['исходная строка']
        else:  # Описание товара 2 типа
            restr = '(' + '|'.join(params2) + ')'
            name, other = rawName.split('; ')
            structName['название'] = name.lower()
            structName.update(dict(map(pars, other.split(', '))))
            addtosearch('название', 'обозначение')
            structName['поиск'] = ' '.join(
                addtosearch('название', 'тип', 'головка', 'обозначение', 'класс прочности', 'материал', 'стандарт')
            )
        p = {}
        for key, value in structName.items():
            if key != 'исходная строка' and key != 'поиск':
                p[key] = value
        outDict = {
            'исходная строка': structName['исходная строка'],
            'поиск': structName['поиск'],
            'параметры': p
        }
        self.structuredData += (outDict,)

        return self.structuredData[self.line_number - 3]

    def output_info(self, number):
        # wb = openpyxl.open('testfile.xlsx')  # Открываем тестовый Excel файл
        self.file.create_sheet(f'Sheet{number}')  # Создаем лист с названием "Sheet1"
        worksheet = self.file[f'Sheet{number}']  # Делаем его активным
        worksheet['B4'] = 'We are writing to B4'  # В указанную ячейку на активном листе пишем все, что в кавычках
        self.file.save('testdel.xlsx')  # Сохраняем измененный файл
